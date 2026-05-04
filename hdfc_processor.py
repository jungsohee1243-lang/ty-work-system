
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =========================
# 설정 부분
# =========================

KEYWORDS_V = [
    "mini PC", "Vibrator", "smart watch", "Earphones",
    "tablet android", "tablet pc", "Speakers", "smartphone",
    "lenovo xiaoxin pad", "IPAD TABLET", "TABLET", "DRAWING BOARD",
    "BLUETOOTH SPEAKER", "BLUETOOTH EARPHONE CLIP",
    "BLUETOOTH EARBUDS", "BLUETOOTH",
]



# ✅ WIRELESS 전용 키워드(정확히 이 문구들만 매칭)
WIRELESS_TERMS = [
    "Wireless Earphones",
    "Wireless Headphones",
    "Wireless Earbuds",
    "Wireless Bluetooth Earphones",
    "Wireless Bluetooth Headphones",
]
# 1차 재분배
WEIGHT_CYCLE = [1.5, 1.6, 1.7, 1.8, 1.9]

# 2차 재분배(목표치 맞출 때) → 0.1kg 단위
STEP_WEIGHT = 0.1

# HS 코드 변환 예외 (V=3 인 경우 변경하지 않을 코드들)
HS_EXCEPT_FOR_V3 = {"900290", "900410", "902920"}

# FTA 적용 HS 코드 리스트
FTA_HS_CODES = {
    "330430", "420292", "630622", "630629", "640219", "640299",
    "640419", "732112", "846729", "847130", "847141", "847150",
    "847160", "847180", "850440", "850760", "850811", "851511",
    "851671", "851762", "851810", "851830", "851840", "852351",
    "852499", "852589", "852691", "852852", "852862", "852869",
    "854231", "854370", "870829", "870894", "870899", "871200",
    "871496", "871499", "900290", "900410", "902920", "910212",
    "910591", "940169", "950450", "950490", "950691",
}


def find_column_name(columns, keyword, startswith=False):
    for c in columns:
        s = str(c)
        if startswith:
            if s.startswith(keyword):
                return c
        else:
            if keyword in s:
                return c
    raise ValueError(f"'{keyword}' 컬럼명을 찾을 수 없습니다.")


def distribute_to_target(df, col_af, target_total):
    w_series = pd.to_numeric(df[col_af], errors="coerce")
    current_total = float(w_series.sum())
    remaining = float(target_total) - current_total

    print(f"[8-1] 현재 AF 합계: {current_total:.3f} kg")
    print(f"[8-2] 목표 AF 합계: {target_total:.3f} kg")
    print(f"[8-3] 추가 필요 중량: {remaining:.3f} kg")

    if remaining <= 0:
        print(" → 이미 목표 이상이므로 분배 없음.")
        return df, current_total, 0.0

    candidates = df.index[pd.to_numeric(df[col_af], errors="coerce") >= 2].tolist()
    if not candidates:
        print(" → 분배 가능한 행 없음.")
        return df, current_total, 0.0

    distributed = 0.0
    i = 0
    n = len(candidates)
    max_loops = 2_000_000

    while remaining >= STEP_WEIGHT - 1e-12 and max_loops > 0:
        idx = candidates[i % n]
        cur = float(w_series.loc[idx])
        new_val = cur + STEP_WEIGHT

        if abs(new_val - 30.0) < 1e-12:
            i += 1
            max_loops -= 1
            continue

        w_series.loc[idx] = new_val
        df.at[idx, col_af] = new_val

        distributed += STEP_WEIGHT
        remaining -= STEP_WEIGHT
        i += 1
        max_loops -= 1

    if remaining > 0:
        print(f"[8-4] 잔여량 {remaining:.3f} kg → 가장 무거운 행에 추가 시도")
        w_series = pd.to_numeric(df[col_af], errors="coerce")
        max_idx = w_series.idxmax()
        cur = float(w_series.loc[max_idx])
        new_val = cur + remaining

        if abs(new_val - 30.0) < 1e-12:
            print(" → 잔여량을 넣으면 30kg이 되어 넣지 않고 종료")
        else:
            df.at[max_idx, col_af] = new_val
            distributed += remaining
            w_series.loc[max_idx] = new_val
            print(f" → {max_idx} 행에 {remaining:.3f} kg 추가 완료")
            remaining = 0.0

    new_total = float(w_series.sum())
    print(f"[8-5] 실제 분배된 총 중량: {distributed:.3f} kg")
    print(f"[8-6] 분배 후 AF 합계: {new_total:.3f} kg")

    return df, new_total, distributed


def process_file(input_path, target_total=None):
    print(f"[1] 파일 읽는 중: {input_path}")
    df = pd.read_excel(input_path)

    print("[2] 컬럼 찾는 중...")
    col_hs    = find_column_name(df.columns, "허용품목코드")
    col_zip   = find_column_name(df.columns, "ZIP CODE")
    col_v     = find_column_name(df.columns, "용도구분")
    col_desc1 = find_column_name(df.columns, "1.DESCRIPTION", startswith=True)
    col_desc2 = find_column_name(df.columns, "2.DESCRIPTION", startswith=True)
    col_af    = find_column_name(df.columns, "Total W/T")
    col_hawb  = find_column_name(df.columns, "HAWB NO")
    col_tel   = find_column_name(df.columns, "C/TEL")
    col_total = find_column_name(df.columns, "总金额")

    w_orig = pd.to_numeric(df[col_af], errors="coerce")
    count_le2_orig = int(((w_orig <= 2) & w_orig.notna()).sum())
    print(f"[3] 원본 AF ≤ 2 : {count_le2_orig} 건")

    print("[4] 허용품목코드 변환...")
    def convert_hs_row(row):
        val_hs = row[col_hs]
        v_val = row[col_v]
        if pd.isna(val_hs):
            return val_hs
        s = str(val_hs).strip()

        if str(v_val).strip() == "3" and s in HS_EXCEPT_FOR_V3:
            return s

        if s.startswith(("1", "2", "30", "90")):
            return "960719"
        return s

    df[col_hs] = df.apply(convert_hs_row, axis=1)

    print("[5] ZIP 4자리 → 5자리...")
    def fix_zip(z):
        if pd.isna(z):
            return z
        s = str(z).strip()
        if len(s) == 4:
            return "0" + s
        return s
    df[col_zip] = df[col_zip].apply(fix_zip)

    print("[6] 용도구분 키워드 조정...")

    # ✅ WIRELESS 전용: V=1 + 1.DESCRIPTION에 WIRELESS 포함 → V=3 변경 (파란색 표시 대상)
    v_before_str = df[col_v].astype(str).str.strip()
    wireless_mask = (v_before_str == "1") & df[col_desc1].astype(str).apply(lambda x: any(term.upper() in str(x).upper() for term in WIRELESS_TERMS))
    df.loc[wireless_mask, col_v] = 3
    rows_v_blue = df.index[wireless_mask].tolist()
    wireless_changed_cnt = int(wireless_mask.sum())
    wireless_ratio_all = (wireless_changed_cnt / len(df) * 100) if len(df) else 0.0
    print(f"   → WIRELESS(지정 문구)로 V=3 변경: {wireless_changed_cnt} 건 ({wireless_ratio_all:.2f}%)")

    def match_v(desc, v):
        if pd.isna(desc) or pd.isna(v):
            return False
        if str(v).strip() != "1":
            return False
        t = str(desc).upper()
        return any(kw.upper() in t for kw in KEYWORDS_V)

    mask_v = df.apply(lambda r: match_v(r[col_desc1], r[col_v]), axis=1)
    df.loc[mask_v, col_v] = 3
    rows_v_red = df.index[mask_v].tolist()
    print(f"   → 키워드로 V=3 변경: {len(rows_v_red)} 건")

    print("[7] AF 2~5 범위 1.5~1.9 재분배...")
    w = pd.to_numeric(df[col_af], errors="coerce")
    mask_range = (w >= 2) & (w <= 5)
    bp_empty = df[col_desc2].isna() | (df[col_desc2].astype(str).str.strip() == "")
    bh_no_elec = ~df[col_desc1].astype(str).str.upper().str.contains("ELECTRIC", na=False)
    mask_target = mask_range & bp_empty & bh_no_elec
    t_idx = df.index[mask_target].tolist()
    print(f"   → 1차 재분배 대상 {len(t_idx)} 건")

    for i, idx in enumerate(t_idx):
        df.at[idx, col_af] = WEIGHT_CYCLE[i % len(WEIGHT_CYCLE)]

    distributed_total = None
    new_total = None
    if target_total is not None:
        print("[8] 0.1kg 단위 목표 중량 분배 시작...")
        df, new_total, distributed_total = distribute_to_target(df, col_af, target_total)
    else:
        print("[8] 목표 총중량 없음 → 2차 분배 없음")

    w_after = pd.to_numeric(df[col_af], errors="coerce")
    count_le2_after = int(((w_after <= 2) & w_after.notna()).sum())
    print(f"[8-7] 분배 후 AF ≤ 2 : {count_le2_after} 건")

    print("[9] 전화번호 중복 + 총금액 합계 ≥150 필터링...")
    tel = df[col_tel].astype(str).str.strip()
    v_str = df[col_v].astype(str).str.strip()
    amt = pd.to_numeric(df[col_total], errors="coerce")

    mask_v1 = (v_str == "1") & tel.notna() & (tel != "")
    df_v1 = pd.DataFrame({"TEL": tel.where(mask_v1), "AMT": amt.where(mask_v1)})
    tel_sum = df_v1.groupby("TEL", dropna=True)["AMT"].sum()
    bad_tels = set(tel_sum[tel_sum >= 150].index.tolist())

    mask_phone_rule = mask_v1 & tel.isin(bad_tels)
    rows_v_orange = df.index[mask_phone_rule].tolist()
    print(f"   → 전화번호 중복 & 총금액합계≥150 행 수: {len(rows_v_orange)} 건")
    print(f"   → 해당 전화번호 수: {len(bad_tels)} 개")

    df.loc[mask_phone_rule, col_v] = 3
    hawb_list = df.loc[mask_phone_rule, col_hawb].astype(str).tolist()

    print("[9-1] FTA 적용건 필터링...")
    v_after_str = df[col_v].astype(str).str.strip()
    hs_str = df[col_hs].astype(str).str.strip()
    amt_after = pd.to_numeric(df[col_total], errors="coerce")

    mask_fta = (v_after_str == "3") & (amt_after >= 150) & hs_str.isin(FTA_HS_CODES)
    fta_hawb_list = df.loc[mask_fta, col_hawb].astype(str).tolist()
    print(f"   → FTA 적용건 행 수: {len(fta_hawb_list)} 건")

    base, _ = os.path.splitext(input_path)
    tmp = base + "_처리값_tmp.xlsx"
    final = base + "_중량조정_최종_v9_wireless.xlsx"

    df.to_excel(tmp, index=False)

    print("[10] 색상 적용 + 메모 시트 생성...")
    wb = load_workbook(tmp)
    ws = wb.active

    header = next(ws.iter_rows(min_row=1, max_row=1))
    col_idx = {cell.value: cell.column for cell in header}
    v_idx = col_idx[col_v]
    af_idx = col_idx[col_af]

    blue   = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")
    red    = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    green  = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
    orange = PatternFill(start_color="FFFF9900", end_color="FFFF9900", fill_type="solid")

    rows_v_blue_excel   = {i + 2 for i in rows_v_blue}
    rows_v_red_excel    = {i + 2 for i in rows_v_red}
    rows_af_excel       = {i + 2 for i in t_idx}
    rows_v_orange_excel = {i + 2 for i in rows_v_orange}

    for r in range(2, ws.max_row + 1):
        if r in rows_v_blue_excel:
            ws.cell(row=r, column=v_idx).fill = blue
        if r in rows_v_red_excel:
            ws.cell(row=r, column=v_idx).fill = red
        if r in rows_af_excel:
            ws.cell(row=r, column=af_idx).fill = green
        if r in rows_v_orange_excel:
            ws.cell(row=r, column=v_idx).fill = orange

    memo = wb.create_sheet("메모")
    memo["A1"] = "AF ≤ 2 (원본)"
    memo["B1"] = count_le2_orig

    memo["A2"] = "목표 총중량(kg)"
    memo["B2"] = "" if target_total is None else float(target_total)

    memo["A3"] = "분배 후 총중량(kg)"
    memo["B3"] = "" if new_total is None else float(new_total)

    memo["A4"] = "2차 분배량 합계(kg)"
    memo["B4"] = "" if distributed_total is None else float(distributed_total)

    memo["A5"] = "AF ≤ 2 (분배 후)"
    memo["B5"] = count_le2_after

    # ✅ WIRELESS 변경 통계(확인용)
    memo["A6"] = "WIRELESS 변경 건수(V=1→3)"
    memo["B6"] = wireless_changed_cnt
    memo["A9"] = "WIRELESS 전체 대비 변경 비율(%)"
    memo["B9"] = round(wireless_ratio_all, 2)

    memo["A7"] = "전화번호 중복 + 총금액합계≥150 행 수"
    memo["B7"] = len(rows_v_orange)

    memo["A8"] = "해당 전화번호 수"
    memo["B8"] = len(bad_tels)

    memo["A10"] = "해당 HAWB NO 리스트"
    row = 11
    for h in hawb_list:
        memo[f"A{row}"] = h
        row += 1

    row += 1
    memo[f"A{row}"] = "FTA적용건 HAWB 리스트"
    memo[f"B{row}"] = len(fta_hawb_list)
    row += 1
    for h in fta_hawb_list:
        memo[f"A{row}"] = h
        row += 1

    wb.save(final)

    try:
        os.remove(tmp)
        print(f"[11] 임시 파일 삭제: {tmp}")
    except Exception as e:
        print(f"[11] 임시 파일 삭제 실패(무시 가능): {e}")

    print(f"완료 → {final}")
    return final


if __name__ == "__main__":
    print("=== HDFC 자동 처리 프로그램 v8 (FTA 적용 포함) ===")
    path = input("파일 경로: ").strip('"').strip()
    if not os.path.isfile(path):
        print("❌ 파일 없음!")
    else:
        t = input("목표 총중량(kg) (없으면 엔터): ").strip()
        target = None
        if t != "":
            try:
                target = float(t)
            except Exception:
                print("❌ 숫자 아님 → 목표 중량 없이 실행")
                target = None
        process_file(path, target_total=target)
